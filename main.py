import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
import time

def geturl(stock_code):
    url = 'https://tw.stock.yahoo.com/q/q?s=' + stock_code
    headers = {'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36'}
    r = requests.get(url, headers = headers, allow_redirects=False)
    # print(r.status_code) # 200表示訪問成功
    # print(r.url)
    soup = BeautifulSoup(r.content, 'html.parser')
    return soup

def getdata(soup):
    stockDic = {}
    table = soup.find_all(text = '成交')[0].parent.parent.parent
    try:
        stockDic['股票名稱'] = table.select('td')[0].select('a')[0].text
        stockDic['時間'] = table.select('td')[1].text
        stockDic['成交價'] = table.select('td')[2].text
        stockDic['買進'] = table.select('td')[3].text
        stockDic['賣出'] = table.select('td')[4].text
        for i in table.select('td')[5].select('font')[0]:
            stockDic['漲跌'] = i.strip()
            break
        stockDic['張數'] = table.select('td')[6].text
        stockDic['昨日收盤價'] = table.select('td')[7].text
        stockDic['開盤價'] = table.select('td')[8].text
        stockDic['最高'] = table.select('td')[9].text
        stockDic['最低'] = table.select('td')[10].text
        stockDic['資料獲取時間'] = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())

    except:
        print('cant get table data')

    # print(stockDic)
    return stockDic


def write_in_excel(stock_data):
    stock_title = []
    stock_value =[]
    excel_name = stock_data['股票名稱']
    row_count = 0

    for key,value in stock_data.items():
        try:
            stock_title.append(key)
            if value.isdigit():
                stock_value.append(int(value))
            else:
                stock_value.append(value)

        except:
            print('Something wrong')

    # 如果目錄已有相關檔案
    try:
        wb = load_workbook(filename = excel_name + '.xlsx')
        ws1 = wb.active
        sheet_range = wb[ws1.title]

        # 確認excel檔中那些rows已經有內容
        for data in sheet_range.rows:
            row_count += 1

        # 將檔案添加在文末
        for row in range(row_count + 1, row_count + 2):
            ws1.append(stock_value)

        # 儲存
        wb.save(filename = excel_name + '.xlsx')

        print('目錄已有相關資料，開始添加' + excel_name + '的今日相關資訊')
    
    except:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = 'Stock Data'

        # 將資料放入xml中
        for row in range(0,1):
            ws1.append(stock_title)
            ws1.append(stock_value)

        # 儲存
        wb.save(filename = excel_name + '.xlsx')

        print('過去沒有相關資料，開始獲取' + excel_name + '的今日相關資訊')

    # 在terminal顯示抓取的cell內容
    max_row = ws1.max_row
    max_column = ws1.max_column
    # 只顯示最新增加的內容，起始row為max_row
    for row in range(max_row, max_row+1):
        for column in range(1, max_column+1):
            cell_obj = ws1.cell(row=row,column=column)
            print(cell_obj.value,end=' | ')
        print('\n') #斷行


    return excel_name, ws1.title




def fixcolumnwidth(excel_name, sheet_name):
    wb = load_workbook(filename = excel_name + '.xlsx')
    sheet_range = wb[sheet_name] # 選擇Stock Data這個Sheet
    col_widths_dict = {}
    
    # 尋找每個column的cell有多少個字，以字數最多的cell為基準放大此column的width
    for row in sheet_range.rows:
        for cell in row:
            #將所有內容置中
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.value:
                # 將每個column的代號column_letter放入dictionary的key，比較完大小後放入value
                col_widths_dict[cell.column_letter] = max((col_widths_dict.get(cell.column_letter, 0), len(str(cell.value))))

    # 調整column width
    for col, value in col_widths_dict.items():
        sheet_range.column_dimensions[col].width = value + 8

    # 調整完後存入原檔案中
    wb.save(filename = excel_name + '.xlsx')

def main():
    try:
        stockID = input('請輸入股票代碼：')
    except ValueError:
        print('股票代碼只可以是整數數字')
    stock_data = getdata(geturl(stockID))
    excel_data_name, sheet_title = write_in_excel(stock_data)
    try:
        fixcolumnwidth(excel_data_name, sheet_title)
    except:
        print('error')

main()